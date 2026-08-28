"""Engagement register → Excel workbook (Phase 3).

Renders the Phase 2 register into ONE stable .xlsx per client/project,
overwritten in place every run so "just export it again after the next
meeting" reflects the full current state.

Hand-edits are expected and must survive regeneration. The contract:

  * Machine columns (text, kind, provenance, …) are regenerated every
    run from the register.
  * Two human-owned columns — Status, Notes — are NEVER generated; they
    are carried forward from the prior workbook, joined by the durable
    record id, falling back to normalized text so a re-extraction that
    mints new ids can't orphan someone's notes.
  * An item that was in the prior workbook but isn't in the new
    register doesn't vanish — it stays, flagged "carried over (last
    seen …)", with its human columns intact.

Safety first: if a prior workbook exists but can't be read (open in
Excel, mid-OneDrive-sync, corrupt) we DO NOT overwrite it — losing
hand-entered triage is unacceptable — we write a dated conflict copy
and return a warning instead.

openpyxl is imported lazily so a missing wheel is a clear API error,
never an import-time crash of the whole backend (mirrors how the
summarizer lazily imports the openai SDK).
"""

from __future__ import annotations

import datetime
import os
import re
import tempfile
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

from services.engagement_service import _norm
from utils.logger import get_logger

logger = get_logger(__name__)

_TERMINAL = {"met", "dropped", "done", "answered"}
_HUMAN_COLS = ("Status", "Notes")
_TAIL_COLS = ("First seen", "Last seen", "# Sessions", "Sessions", "Detected")

# Per register key: sheet title + the machine content columns. The
# first content column is the record's identity/primary text (used for
# the normalized-text carry-forward fallback).
SHEETS: List[dict] = [
    {"key": "requirements", "title": "Requirements", "cols": [
        ("Requirement", lambda r: r.get("text", "")),
        ("Kind", lambda r: r.get("kind", "")),
        ("Source", lambda r: r.get("source", "")),
    ]},
    {"key": "decisions", "title": "Decisions", "cols": [
        ("Decision", lambda r: r.get("title", "")),
        ("Decided", lambda r: r.get("decided", "")),
        ("Rationale", lambda r: r.get("rationale", "")),
        ("Alternatives", lambda r: r.get("alternatives", "")),
        ("Owner", lambda r: r.get("owner", "")),
        ("Impact", lambda r: r.get("impact", "")),
        ("Source", lambda r: r.get("source", "")),
    ]},
    {"key": "action_items", "title": "Action Items", "cols": [
        ("Action", lambda r: r.get("text", "")),
        ("Owner", lambda r: r.get("owner", "")),
        ("Due", lambda r: r.get("due") or ""),
        ("Source", lambda r: r.get("source", "")),
    ]},
    {"key": "open_questions", "title": "Open Questions", "cols": [
        ("Question", lambda r: r.get("text", "")),
        ("Source", lambda r: r.get("source", "")),
    ]},
    # Defect register. Ref leads because that is what triage calls the
    # row by, and the sheet is read in the meeting it came from.
    {"key": "defects", "title": "Defects", "cols": [
        ("Defect", lambda r: r.get("title", "")),
        ("Ref", lambda r: r.get("ref", "")),
        ("Severity", lambda r: r.get("severity", "")),
        ("Owner", lambda r: r.get("owner", "")),
        ("Due", lambda r: r.get("due") or ""),
        ("Disposition", lambda r: r.get("disposition", "")),
        ("Source", lambda r: r.get("source", "")),
    ]},
]


def _headers(spec: dict) -> List[str]:
    return (["Ref ID"] + [h for h, _ in spec["cols"]]
            + list(_TAIL_COLS) + list(_HUMAN_COLS))


def _provenance(rec: dict) -> Tuple[str, str, int, str]:
    occ = rec.get("occurrences") or []
    first = occ[0]["at"] if occ else ""
    last = occ[-1]["at"] if occ else ""
    sessions = "; ".join(
        f"{o.get('display_name') or o.get('session_id')} "
        f"({(o.get('at') or '')[:10]})"
        for o in occ
    )
    return first, last, len(occ), sessions


def _safe_name(name: str) -> str:
    name = re.sub(r"[^A-Za-z0-9._ &-]+", "_", name).strip()
    return name or "Engagement Register"


# ---- prior-workbook readback (for carry-forward) -------------------

def _read_prior(path: Path) -> Optional[dict]:
    """Return {'generated_at': str, 'sheets': {title: [rowdict,...]}}
    or None if there's nothing to read. Raises on a file that EXISTS
    but can't be opened — the caller turns that into a non-destructive
    conflict write, never an overwrite."""
    if not path.exists():
        return None
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out: dict = {"generated_at": "", "sheets": {}}
        if "Overview" in wb.sheetnames:
            for row in wb["Overview"].iter_rows(values_only=True):
                if row and str(row[0] or "").strip().lower() == "generated at":
                    out["generated_at"] = str(row[1] or "")
                    break
        for spec in SHEETS:
            title = spec["title"]
            if title not in wb.sheetnames:
                continue
            ws = wb[title]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c or "") for c in rows[0]]
            recs = []
            for raw in rows[1:]:
                d = {header[i]: ("" if raw[i] is None else str(raw[i]))
                     for i in range(min(len(header), len(raw)))}
                if not any(v.strip() for v in d.values()):
                    continue
                recs.append(d)
            out["sheets"][title] = recs
        return out
    finally:
        wb.close()


# ---- workbook build -------------------------------------------------

def _build_workbook(register: dict, prior: Optional[dict]):
    """Pure-ish: register (+ optional prior readback) → an in-memory
    openpyxl Workbook. No filesystem writes here so it's unit-testable.
    Returns (workbook, changes_summary:dict)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(bold=True, color="FFFFFF")
    human_fill = PatternFill("solid", fgColor="FFF3C4")  # "yours" = amber
    human_font = Font(bold=True, color="7A5D00")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    wb.remove(wb.active)

    prior_sheets = (prior or {}).get("sheets", {})
    prior_gen = (prior or {}).get("generated_at", "")
    changes: Dict[str, Dict[str, list]] = {}

    # Overview ------------------------------------------------------
    ov = wb.create_sheet("Overview")
    counts = register.get("counts", {})
    meta = [
        ("Client", register.get("client", "")),
        ("Project", register.get("project", "") or "(all projects)"),
        ("Generated at", register.get("generated_at", "")),
        ("Sessions", register.get("session_count", 0)),
        ("Open requirements", counts.get("open_requirements", 0)),
        ("Decisions", counts.get("decisions", 0)),
        ("Open action items", counts.get("open_action_items", 0)),
        ("Open questions", counts.get("open_questions", 0)),
        ("Previous export", prior_gen or "(first export)"),
    ]
    for i, (k, v) in enumerate(meta, 1):
        ov.cell(i, 1, k).font = Font(bold=True)
        ov.cell(i, 2, v)
    ov.column_dimensions["A"].width = 22
    ov.column_dimensions["B"].width = 48

    for spec in SHEETS:
        key, title, cols = spec["key"], spec["title"], spec["cols"]
        headers = _headers(spec)
        ws = wb.create_sheet(title)
        ws.append(headers)
        for ci, _h in enumerate(headers, 1):
            c = ws.cell(1, ci)
            is_human = _h in _HUMAN_COLS
            c.fill = human_fill if is_human else hdr_fill
            c.font = human_font if is_human else hdr_font
        ws.freeze_panes = "A2"

        prior_rows = prior_sheets.get(title, [])
        prior_by_id = {r.get("Ref ID", ""): r
                       for r in prior_rows if r.get("Ref ID")}
        # primary text is the first content column → header index 1
        prim_hdr = headers[1]
        prior_by_norm = {_norm(r.get(prim_hdr, "")): r
                         for r in prior_rows if r.get(prim_hdr)}

        recs = register.get(key) or []
        new_ids, new_norms = set(), set()
        out_rows: List[Tuple[list, bool]] = []  # (row, is_carried_over)
        new_items, resolved_items = [], []

        for rec in recs:
            rid = rec.get("id", "")
            primary = cols[0][1](rec)
            nkey = _norm(primary)
            new_ids.add(rid)
            new_norms.add(nkey)
            carried = prior_by_id.get(rid) or prior_by_norm.get(nkey)
            status = (carried or {}).get("Status", "")
            notes = (carried or {}).get("Notes", "")
            first, last, n, sess = _provenance(rec)
            row = [rid] + [fn(rec) for _h, fn in cols] + [
                first, last, n, sess, "current", status, notes]
            out_rows.append((row, False))
            if carried is None:
                new_items.append(primary)
            if (rec.get("status") or "").lower() in _TERMINAL:
                resolved_items.append(primary)

        # Prior rows no longer in the register → keep, don't lose edits.
        gone = []
        for pr in prior_rows:
            pid = pr.get("Ref ID", "")
            pnorm = _norm(pr.get(prim_hdr, ""))
            if pid in new_ids or (pnorm and pnorm in new_norms):
                continue
            prev_detected = pr.get("Detected", "") or ""
            if prev_detected.startswith("carried over"):
                detected = prev_detected  # already stamped; keep date
            else:
                seen = (prior_gen or "")[:10] or "?"
                detected = f"carried over (last seen {seen})"
                gone.append(pr.get(prim_hdr, ""))
            row = [pr.get(h, "") for h in headers]
            # force the regenerated Detected, keep human cols as-is
            row[headers.index("Detected")] = detected
            out_rows.append((row, True))

        # Open/current first; resolved + carried-over sink (history kept).
        def sort_key(item):
            row, carried = item
            st = str(row[headers.index("Status")] or "").lower()
            terminal = st in _TERMINAL
            return (carried, terminal)
        out_rows.sort(key=sort_key)

        for row, _carried in out_rows:
            ws.append(row)

        # Column widths + wrap the free-text primary/notes columns.
        for ci, h in enumerate(headers, 1):
            letter = get_column_letter(ci)
            if h in ("Requirement", "Decision", "Action", "Question",
                     "Decided", "Rationale", "Sessions", "Notes"):
                ws.column_dimensions[letter].width = 42
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, ci).alignment = wrap
            elif h == "Ref ID":
                ws.column_dimensions[letter].width = 14
            else:
                ws.column_dimensions[letter].width = 18

        changes[title] = {
            "new": new_items,
            "gone": gone,
            "resolved": resolved_items,
        }

    # Changes-since-last-export ------------------------------------
    cs = wb.create_sheet("Changes since last export")
    since = prior_gen or "(first export — everything is new)"
    cs.append(["Since", since])
    cs.append([])
    cs.append(["Type", "Change", "Item"])
    for ci in range(1, 4):
        c = cs.cell(3, ci)
        c.fill = hdr_fill
        c.font = hdr_font
    cs.freeze_panes = "A4"
    for spec in SHEETS:
        ch = changes.get(spec["title"], {})
        for label, bucket in (("New", "new"),
                              ("No longer detected", "gone"),
                              ("Resolved", "resolved")):
            for item in ch.get(bucket, []):
                cs.append([spec["title"], label, item])
    cs.column_dimensions["A"].width = 16
    cs.column_dimensions["B"].width = 22
    cs.column_dimensions["C"].width = 70

    return wb, changes


# ---- public entrypoint ---------------------------------------------

def export_register_workbook(
    register: dict, dest_dir, filename: str
) -> dict:
    """Render `register` to dest_dir/filename, carrying human columns
    forward from any existing workbook there. Returns
    {"path": str, "warning": Optional[str]}. Never raises for the
    'file is open / locked' case — that yields a dated conflict copy
    plus a warning so hand-entered edits are never destroyed."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        ) from e

    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    dest = dest_dir / _safe_name(filename)
    warning: Optional[str] = None

    # Read prior for carry-forward. If it exists but won't open, the
    # only safe move is to NOT touch it and write a conflict copy —
    # overwriting would silently destroy hand-entered Status/Notes.
    prior = None
    prior_unreadable = False
    try:
        prior = _read_prior(dest)
    except Exception as e:
        prior_unreadable = True
        warning = (
            f"Existing workbook could not be read ({e}); it's likely open "
            f"in Excel or still syncing. To avoid destroying any hand-"
            f"entered Status/Notes, the update was written to a separate "
            f"dated copy — close the original and re-export to merge."
        )
        logger.warning(f"prior register unreadable, not overwriting: {e}")

    wb, _changes = _build_workbook(register, prior)

    if prior_unreadable:
        stamp = datetime.datetime.now().strftime("%Y-%m-%d %H%M")
        target = dest.with_name(f"{dest.stem} (conflicted {stamp}).xlsx")
        wb.save(target)
        return {"path": str(target), "warning": warning}

    # Atomic replace so a reader never sees a half-written file.
    tmp = None
    try:
        fd, tmp_name = tempfile.mkstemp(
            dir=dest_dir, prefix=".er_", suffix=".xlsx")
        os.close(fd)
        tmp = Path(tmp_name)
        wb.save(tmp)
        os.replace(tmp, dest)
        tmp = None
        return {"path": str(dest), "warning": None}
    except PermissionError:
        # dest is open/locked (Excel, OneDrive). Don't fail — write a
        # dated copy alongside and tell the caller.
        stamp = datetime.datetime.now().strftime("%Y-%m-%d %H%M")
        target = dest.with_name(f"{dest.stem} (conflicted {stamp}).xlsx")
        wb.save(target)
        return {
            "path": str(target),
            "warning": (
                f"{dest.name} is open or locked; wrote {target.name} "
                f"instead. Close it and re-export to update in place."
            ),
        }
    finally:
        if tmp is not None and tmp.exists():
            tmp.unlink(missing_ok=True)
